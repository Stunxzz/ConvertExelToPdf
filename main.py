from openpyxl import load_workbook
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from pdf2image import convert_from_path

def excel_to_pdf(excel_file, pdf_file):
    # Load the Excel workbook
    wb = load_workbook(excel_file)
    # Get the active worksheet
    ws = wb.active
    
    c = canvas.Canvas(pdf_file, pagesize=letter)

    y = 750

    for row in ws.iter_rows():
        # Set initial x-coordinate
        x = 50
        for cell in row:
            # Write cell value to PDF
            c.drawString(x, y, str(cell.value))
            x += 100  # Increase x-coordinate for the next cell
        y -= 12  # Decrease y-coordinate for the next row

    c.save()


def convert_to_image(pdf_file):
    # Convert PDF pages to images
    pages = convert_from_path(pdf_file)

    # Save each page as an image
    for i, page in enumerate(pages):
        image_path = f'page_{i + 1}.jpg'  # Change the extension to '.png' or '.jpeg' if desired
        page.save(image_path, 'JPEG')

excel_file = 'template.xlsx'
pdf_file = 'template.pdf'
excel_to_pdf(excel_file, pdf_file)


convert_to_image(pdf_file)